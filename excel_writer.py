from pathlib import Path
from difflib import SequenceMatcher
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 실제 사용 중인 템플릿 컬럼 순서 (고정)
COLUMNS = [
    "TC_ID",
    "카테고리",
    "테스트 제목",
    "테스트 목적",
    "사전 조건",
    "입력값",
    "테스트 절차",
    "기대결과",
    "검토",
    "P/F",
    "비고",
]

HEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_FONT = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMN_WIDTHS = {
    "TC_ID": 14,
    "카테고리": 14,
    "테스트 제목": 28,
    "테스트 목적": 26,
    "사전 조건": 22,
    "입력값": 18,
    "테스트 절차": 40,
    "기대결과": 30,
    "검토": 10,
    "P/F": 8,
    "비고": 20,
}


def create_template(path: str) -> None:
    """헤더만 있는 새 테스트케이스 Excel 파일 생성. 이미 파일이 있으면 아무 것도 하지 않음."""
    file_path = Path(path)
    if file_path.exists():
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "TestCases"

    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS[col_name]

    ws.freeze_panes = "A2"
    file_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(file_path)


def _next_tc_id(ws, bug_id: str) -> str:
    """같은 bug_id 접두사를 가진 TC_ID 중 가장 큰 순번의 다음 번호 반환. 예: TC_298_01, TC_298_02 ..."""
    prefix = f"TC_{bug_id}_"
    max_seq = 0
    for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
        value = row[0]
        if value and str(value).startswith(prefix):
            try:
                seq = int(str(value).split("_")[-1])
                max_seq = max(max_seq, seq)
            except ValueError:
                continue
    return f"{prefix}{max_seq + 1:02d}"


def append_test_case(path: str, bug_id: str, tc_data: dict) -> str:
    """
    AI가 생성한 테스트케이스 dict를 Excel에 한 행 추가.

    tc_data 예상 키:
        category, title, purpose, precondition, input_value,
        steps (list[str] 권장, str도 허용), expected

    '검토' / 'P/F' / '비고'는 QA가 직접 채우는 영역이라 빈 칸으로 남김.

    반환값: 새로 채번된 TC_ID
    """
    create_template(path)  # 파일이 없으면 생성
    wb = load_workbook(path)
    ws = wb["TestCases"] if "TestCases" in wb.sheetnames else wb.active

    tc_id = _next_tc_id(ws, bug_id)

    steps = tc_data.get("steps", [])
    if isinstance(steps, list):
        steps_text = "\n".join(f"{i + 1}. {s}" for i, s in enumerate(steps))
    else:
        steps_text = str(steps)

    row_values = [
        tc_id,
        tc_data.get("category", ""),
        tc_data.get("title", ""),
        tc_data.get("purpose", ""),
        tc_data.get("precondition", ""),
        tc_data.get("input_value", ""),
        steps_text,
        tc_data.get("expected", ""),
        "",  # 검토
        "",  # P/F
        "",  # 비고
    ]

    next_row = ws.max_row + 1
    for col_idx, value in enumerate(row_values, start=1):
        cell = ws.cell(row=next_row, column=col_idx, value=value)
        cell.alignment = WRAP
        cell.border = BORDER

    try:
        wb.save(path)
    except PermissionError as e:
        raise PermissionError(
            f"'{path}' 파일에 저장할 수 없습니다. "
            "Excel에서 이 파일을 열어두신 상태라면 닫고 다시 시도해주세요."
        ) from e

    return tc_id


def _similarity(a: str, b: str) -> float:
    """0.0~1.0 사이 텍스트 유사도 (한글 포함 문자 단위 비교)."""
    return SequenceMatcher(None, a or "", b or "").ratio()


def find_similar_test_cases(path: str, title: str, purpose: str = "", threshold: float = 0.6) -> list:
    """
    기존 Excel에 이미 저장된 테스트케이스들과 title(+purpose)을 비교해서
    유사도가 threshold 이상인 것들을 찾아 반환.

    threshold: 0.6 = 문자 기준 60% 이상 비슷하면 후보로 봄 (완전 동일이 1.0)

    반환: [{"tc_id", "category", "title", "similarity"}, ...] 유사도 내림차순.
          파일이 없거나 비교 대상이 없으면 빈 리스트.
    """
    file_path = Path(path)
    if not file_path.exists():
        return []

    wb = load_workbook(path)
    ws = wb["TestCases"] if "TestCases" in wb.sheetnames else wb.active

    matches = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        existing_tc_id = row[0]
        existing_category = row[1] if len(row) > 1 else ""
        existing_title = row[2] if len(row) > 2 else ""
        existing_purpose = row[3] if len(row) > 3 else ""
        if not existing_title:
            continue

        title_sim = _similarity(title, str(existing_title))
        purpose_sim = _similarity(purpose, str(existing_purpose)) if purpose else 0.0
        # 제목이 비슷하거나, 제목은 좀 다른데 목적(purpose)이 거의 같으면(재검증하는 관점 차이 등) 후보로 봄
        combined = max(title_sim, purpose_sim)

        if combined >= threshold:
            matches.append({
                "tc_id": existing_tc_id,
                "category": existing_category,
                "title": existing_title,
                "similarity": round(combined, 2),
            })

    matches.sort(key=lambda m: m["similarity"], reverse=True)
    return matches


if __name__ == "__main__":
    # AI 호출 없이 모듈 자체 동작만 확인하는 샘플 (298번 버그 기준)
    sample_path = "testcase_sample.xlsx"
    sample_tc = {
        "category": "도메인 관리",
        "title": "도메인명 공백 입력 시 저장 방지 확인",
        "purpose": "도메인명에 공백만 입력했을 때 trim 처리 및 저장 방지 로직이 정상 동작하는지 검증",
        "precondition": "도메인 관리 화면 진입, 신규 도메인 생성 상태",
        "input_value": "도메인명 = \" \" (공백 1개 이상)",
        "steps": [
            "도메인 목록에서 빈 영역 우클릭",
            "'새 도메인 만들기' 선택",
            "도메인명 입력란에 공백만 입력",
            "저장 또는 Tab 키로 포커스 이동",
        ],
        "expected": "trim 결과가 빈 문자열이므로 저장이 차단되고 경고 메시지가 표시되어야 함",
    }
    new_id = append_test_case(sample_path, "298", sample_tc)
    print(f"생성된 TC_ID: {new_id} -> {sample_path}")

    # 중복 trim 케이스도 하나 더 추가해서 채번이 잘 되는지 확인
    sample_tc2 = {
        "category": "도메인 관리",
        "title": "앞뒤 공백 포함 도메인명 중복 등록 방지 확인",
        "purpose": "기존 도메인명과 공백을 제외하면 동일한 이름을 재등록할 수 없는지 검증",
        "precondition": "도메인 'user'가 이미 생성되어 있음",
        "input_value": "도메인명 = \" user \" (앞뒤 공백 포함)",
        "steps": [
            "새 도메인 생성 화면 진입",
            "도메인명에 ' user ' 입력",
            "저장",
        ],
        "expected": "trim 후 기존 'user'와 동일 문자열로 판단되어 중복 오류가 표시되고 저장이 차단되어야 함",
    }
    new_id2 = append_test_case(sample_path, "298", sample_tc2)
    print(f"생성된 TC_ID: {new_id2} -> {sample_path}")